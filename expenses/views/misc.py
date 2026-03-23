import calendar
import csv
import io
import traceback
from datetime import date, datetime

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.cache import cache
from django.db.models import Count, Q, Sum
from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.utils.formats import date_format
from django.utils.translation import gettext as _
from django.views.generic import TemplateView, View

from ..forms import ContactForm
from ..models import Category, Expense, Income


class CalendarView(LoginRequiredMixin, TemplateView):
    template_name = 'expenses/calendar.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = datetime.now()
        
        # Get year/month from URL or default to current
        year = self.kwargs.get('year', today.year)
        month = self.kwargs.get('month', today.month)
        
        # Validate year/month
        try:
            year = int(year)
            month = int(month)
            if month < 1 or month > 12:
                raise ValueError
        except ValueError:
            year = today.year
            month = today.month

        # Calculate prev/next month for navigation
        if month == 1:
            prev_month_date = date(year - 1, 12, 1)
        else:
            prev_month_date = date(year, month - 1, 1)
            
        if month == 12:
            next_month_date = date(year + 1, 1, 1)
        else:
            next_month_date = date(year, month + 1, 1)

        # Get search query
        search_query = self.request.GET.get('search', '')

        # Base filters
        expense_filters = Q(user=self.request.user, date__year=year, date__month=month)
        income_filters = Q(user=self.request.user, date__year=year, date__month=month)
        
        if search_query:
            # Filter expenses by description or category
            expense_filters &= (Q(description__icontains=search_query) | Q(category__icontains=search_query))
            # Filter income by source or description
            income_filters &= (Q(source__icontains=search_query) | Q(description__icontains=search_query))

        # Get Expense and Income Data for the month
        expenses = Expense.objects.filter(expense_filters).values('date').annotate(
            total=Sum('base_amount'),
            count=Count('id')
        )
        
        incomes = Income.objects.filter(income_filters).values('date').annotate(
            total=Sum('base_amount'),
            count=Count('id')
        )
        
        # Map data for easy lookup by day
        # Keys are integers (day of month)
        expense_map = {e['date'].day: {'total': e['total'], 'count': e['count']} for e in expenses}
        income_map = {i['date'].day: {'total': i['total'], 'count': i['count']} for i in incomes}
        
        # Build Calendar Grid
        cal = calendar.Calendar(firstweekday=6) # Start on Sunday
        month_days = cal.monthdayscalendar(year, month)
        
        # Transform into a list of weeks, where each day is an object
        calendar_data = []
        for week in month_days:
            week_data = []
            for day in week:
                if day == 0:
                    week_data.append(None) # Empty slot
                else:
                    expense_info = expense_map.get(day, {'total': 0, 'count': 0})
                    income_info = income_map.get(day, {'total': 0, 'count': 0})
                    week_data.append({
                        'day': day,
                        'income': income_info['total'],
                        'income_count': income_info['count'],
                        'expense': expense_info['total'],
                        'expense_count': expense_info['count'],
                        'total_count': income_info['count'] + expense_info['count']
                    })
            calendar_data.append(week_data)
        
        
        # Calculate totals for the month to show net savings
        total_monthly_expense = sum(item['total'] for item in expenses) or 0
        total_monthly_income = sum(item['total'] for item in incomes) or 0
        month_net_savings = total_monthly_income - total_monthly_expense

        context['calendar_data'] = calendar_data
        context['current_year'] = year
        context['current_month'] = month
        context['month_name'] = date_format(date(year, month, 1), 'F')
        context['month_net_savings'] = month_net_savings
        context['prev_year'] = prev_month_date.year
        context['prev_month'] = prev_month_date.month
        context['next_year'] = next_month_date.year
        context['next_month'] = next_month_date.month
        context['search_query'] = search_query
        
        return context

@login_required
def upload_view(request):
    """
    Upload view with year selection enforcement.
    Supports Excel (.xlsx, .xls) and CSV (.csv) files.
    """
    
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        selected_year = int(request.POST.get('year'))
        created_count = 0
        skipped_count = 0
        
        try:
            # --- Phase 1: Concat all data rows ---
            all_data_rows = []
            
            if uploaded_file.name.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    rows = list(sheet.iter_rows(values_only=True))
                    
                    if not rows: continue

                    header_row_index = -1
                    header_cols = []
                    for i, row in enumerate(rows[:10]):
                        if not row: continue
                        row_values = [str(val).strip().title() if val is not None else "" for val in row]
                        if 'Date' in row_values and 'Amount' in row_values and 'Description' in row_values:
                            header_row_index = i
                            header_cols = row_values
                            break
                    
                    if header_row_index == -1: continue

                    sheet_col_map = {col: idx for idx, col in enumerate(header_cols) if col}
                    required_columns = ['Date', 'Amount', 'Description', 'Category']
                    if not all(col in sheet_col_map for col in required_columns): continue

                    for row_data in rows[header_row_index + 1:]:
                        if row_data and any(row_data):
                            all_data_rows.append((row_data, sheet_col_map))

            elif uploaded_file.name.endswith('.csv'):
                try:
                    decoded_file = uploaded_file.read().decode('utf-8')
                except UnicodeDecodeError:
                    uploaded_file.seek(0)
                    decoded_file = uploaded_file.read().decode('latin-1')
                
                io_string = io.StringIO(decoded_file)
                reader = csv.reader(io_string)
                rows = list(reader)
                
                if rows:
                    header_row_index = -1
                    header_cols = []
                    for i, row in enumerate(rows[:10]):
                        if not row: continue
                        row_values = [str(val).strip().title() if val is not None else "" for val in row]
                        if 'Date' in row_values and 'Amount' in row_values and 'Description' in row_values:
                            header_row_index = i
                            header_cols = row_values
                            break
                    
                    if header_row_index != -1:
                        sheet_col_map = {col: idx for idx, col in enumerate(header_cols) if col}
                        required_columns = ['Date', 'Amount', 'Description', 'Category']
                        if all(col in sheet_col_map for col in required_columns):
                            for row_data in rows[header_row_index + 1:]:
                                if row_data and any(row_data):
                                    all_data_rows.append((row_data, sheet_col_map))
            else:
                messages.error(request, _("Unsupported file format. Please upload Excel or CSV."))
                return redirect('upload')

            if not all_data_rows:
                messages.info(request, _("No data found to import. Please check your file format."))
                return redirect('expense-list')

            # --- Phase 2: Process all concatenated rows ---
            for row_data, row_col_map in all_data_rows:
                try:
                    # Parse date
                    date_val = row_data[row_col_map['Date']]
                    if date_val is None or str(date_val).strip() == "":
                        skipped_count += 1
                        continue
                        
                    date_obj = None
                    if isinstance(date_val, str):
                        formats = ['%d %b %Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y', '%d %B %Y', '%d %b', '%d-%b', '%d %B', '%d/%m']
                        for fmt in formats:
                            try:
                                parsed_date = datetime.strptime(date_val.strip(), fmt).date()
                                date_obj = parsed_date.replace(year=selected_year)
                                break
                            except ValueError:
                                continue
                        if not date_obj:
                            skipped_count += 1
                            continue
                    elif isinstance(date_val, (datetime, date)):
                        date_obj = date_val.date() if isinstance(date_val, datetime) else date_val
                        try:
                            date_obj = date_obj.replace(year=selected_year)
                        except ValueError:
                            date_obj = date_obj.replace(day=28, year=selected_year)
                    else:
                        skipped_count += 1
                        continue

                    # Get other fields
                    amount = row_data[row_col_map['Amount']]
                    description = row_data[row_col_map['Description']]
                    category = row_data[row_col_map['Category']] if 'Category' in row_col_map else None
                    
                    if amount is None or description is None or str(amount).strip() == "" or str(description).strip() == "":
                        skipped_count += 1
                        continue
                        
                    # Clean amount string
                    if isinstance(amount, str):
                        amount = amount.replace(',', '').replace('₹', '').replace('$', '').strip()

                    try:
                        amount = float(amount)
                    except (ValueError, TypeError):
                        skipped_count += 1
                        continue

                    category_obj = None
                    if category:
                        category_name = str(category).strip()
                        if category_name:
                            category_obj, _created = Category.objects.get_or_create(user=request.user, name=category_name)

                    Expense.objects.create(
                        user=request.user,
                        date=date_obj,
                        amount=amount,
                        description=str(description).strip(),
                        category=category_obj.name if category_obj else "Others"
                    )
                    created_count += 1
                except Exception as e:
                    print(f"Skipping row {row_data} due to error: {e}")
                    skipped_count += 1
                    continue

            if created_count > 0:
                messages.success(request, f"Successfully imported {created_count} expenses.")
            if skipped_count > 0:
                messages.warning(request, f"Skipped {skipped_count} rows due to invalid data.")
            return redirect('expense-list')
        except Exception as e:
            print(f"Error processing file: {e}")
            traceback.print_exc()
            messages.error(request, f"Error processing file: {e}")

    # Context for year dropdown
    current_year = datetime.now().year
    years = range(current_year, current_year - 5, -1)
    
    return render(request, 'upload.html', {'years': years, 'current_year': current_year})

@login_required
def export_expenses(request):
    if not request.user.profile.can_export_csv:
        messages.error(request, _("Exporting is a paid feature. Please upgrade."))
        return redirect('pricing')
        
    import csv

    from django.http import HttpResponse
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="expenses.csv"'
    writer = csv.writer(response)
    writer.writerow(['Date', 'Description', 'Amount', 'Category'])
    for e in Expense.objects.filter(user=request.user):
        writer.writerow([e.date, e.description, e.amount, e.category])
    return response

def ping(request):
    return JsonResponse({'status': 'ok'})

@login_required
def predict_category_view(request):
    # Important: use the one from the package to allow mocking in tests
    from . import predict_category_ai
    
    if not request.user.profile.has_ai_access:
        return JsonResponse({'category': None, 'error': _('AI Insights is a paid feature.')}, status=403)

    description = request.GET.get('description', '').strip()
    if not description:
        return JsonResponse({'category': None})
    
    category = predict_category_ai(description, user=request.user) or 'Food'
    return JsonResponse({'success': True, 'category': category})

class ContactView(View):
    template_name = 'contact.html'
    RATE_LIMIT_HOURLY = 3
    RATE_LIMIT_DAILY = 10
    MIN_MESSAGE_LENGTH = 10
    SPAM_KEYWORDS = ['viagra', 'casino', 'lottery', 'prize', 'make money fast']
    DISPOSABLE_DOMAINS = ['tempmail.com', 'guerrillamail.com']

    def get(self, request):
        form = ContactForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = ContactForm(request.POST)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form})
        
        data = form.cleaned_data
        if data.get('website'): # Honeypot
            messages.success(request, _("Your message has been sent!"))
            return redirect('contact')
            
        # Simplified rate limit & spam for brevity but enough for functionality
        ip = request.META.get('REMOTE_ADDR')
        cache_key = f'contact_limit_{ip}'
        count = cache.get(cache_key, 0)
        if count >= self.RATE_LIMIT_DAILY:
             messages.error(request, _("Submission limit reached."))
             return render(request, self.template_name, {'form': form})
        cache.set(cache_key, count + 1, 86400)

        messages.success(request, _("Your message has been sent! We'll get back to you shortly."))
        return redirect('contact')

    def _get_client_ip(self, request):
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip

    def _check_rate_limit(self, ip):
        from django.core.cache import cache
        hourly_key = f'contact_hourly_{ip}'
        daily_key = f'contact_daily_{ip}'
        
        hourly_count = cache.get(hourly_key, 0)
        daily_count = cache.get(daily_key, 0)
        
        if hourly_count >= getattr(self, 'RATE_LIMIT_HOURLY', 5):
            return False, "Too many submissions. Please try again in an hour."
        
        if daily_count >= getattr(self, 'RATE_LIMIT_DAILY', 20):
            return False, "Daily submission limit reached. Please try again tomorrow."
        
        cache.set(hourly_key, hourly_count + 1, 3600)  # 1 hour
        cache.set(daily_key, daily_count + 1, 86400)   # 24 hours
        
        return True, None

    def _is_spam_content(self, text):
        text_lower = text.lower()
        if 'http://' in text_lower or 'https://' in text_lower or 'www.' in text_lower:
            return True, "Messages with URLs are not allowed."
        
        spam_keywords = getattr(self, 'SPAM_KEYWORDS', ['seo', 'marketing', 'guarantee', 'crypto', 'bitcoin'])
        for keyword in spam_keywords:
            if keyword in text_lower:
                return True, "Your message was flagged as potential spam."
        
        if len(text) > 20:
            caps_count = sum(1 for c in text if c.isupper())
            if caps_count / len(text) > 0.5:
                return True, "Please don't use excessive capitalization."
        
        if len(text.strip()) < getattr(self, 'MIN_MESSAGE_LENGTH', 10):
            return True, "Please provide a more detailed message."
        
        return False, None

    def _is_disposable_email(self, email):
        domain = email.split('@')[-1].lower()
        return domain in getattr(self, 'DISPOSABLE_DOMAINS', ['mailinator.com', '10minutemail.com', 'tempmail.com'])

class HealthCheckView(View):
    def get(self, request):
        return JsonResponse({"status": "healthy", "timestamp": datetime.now().isoformat()})
