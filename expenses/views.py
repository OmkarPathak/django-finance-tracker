from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.db import IntegrityError
import csv
from django.forms import modelformset_factory
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth import login, logout
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.views import generic
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView, View
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Q
from django.http import JsonResponse, HttpResponse
import json
from django.utils import timezone
from datetime import datetime, date, timedelta
import calendar

from .models import Expense, Category, Income, RecurringTransaction, UserProfile, SubscriptionPlan, Notification, SIPInvestment, Tag, FilterPreset
from finance_tracker.ai_utils import predict_category_ai
from .forms import ExpenseForm, IncomeForm, RecurringTransactionForm, ProfileUpdateForm, CustomSignupForm, ContactForm, SIPForm, TagForm
from allauth.socialaccount.models import SocialAccount
import openpyxl
import requests
import traceback
from django.core.management import call_command
from allauth.account.models import EmailAddress
from django.core.mail import send_mail
from django.conf import settings
from django.core.cache import cache
from django.db.models.functions import TruncMonth, TruncDay
from django.utils.html import mark_safe, escape, format_html, format_html_join


def create_category_ajax(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            name = data.get('name', '').strip()
            
            if not name:
                return JsonResponse({'success': False, 'error': 'Category name cannot be empty.'}, status=400)
            
            # Check Limits
            current_count = Category.objects.filter(user=request.user).count()
            limit = 5 # Free
            if request.user.profile.is_plus:
                limit = 10
            if request.user.profile.is_pro:
                limit = float('inf')

            if current_count >= limit:
                 return JsonResponse({'success': False, 'error': f'Category limit reached ({limit}). Please upgrade.'}, status=403)

            category = Category.objects.create(user=request.user, name=name)
            return JsonResponse({'success': True, 'id': category.id, 'name': category.name})
            
        except IntegrityError:
            return JsonResponse({'success': False, 'error': 'This category already exists.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
            
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)


def resend_verification_email(request):
    """
    AJAX view to resend verification email.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')
            
            # If email is not in body, try to get from logged in user
            if not email and request.user.is_authenticated:
                email = request.user.email
            
            # Fallback: Check allauth session key
            if not email:
                email = request.session.get('account_email')

            if not email:
                return JsonResponse({'success': False, 'error': 'Email is missing.'}, status=400)
            
            try:
                # Case-insensitive lookup just in case
                email_address = EmailAddress.objects.filter(email__iexact=email).first()
                if not email_address:
                     return JsonResponse({'success': False, 'error': f'Email {email} not found in system.'}, status=404)
                
                # Check if already verified
                if email_address.verified:
                    return JsonResponse({'success': True, 'message': 'Email already verified.'})

                email_address.send_confirmation(request)
                return JsonResponse({'success': True, 'message': 'Verification email sent!'})

            except Exception as e:
                # Log the actual error for debugging
                
                print(traceback.format_exc())
                return JsonResponse({'success': False, 'error': f'Send failed: {str(e)}'}, status=500)
                
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Server Error: {str(e)}'}, status=500)
            
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

def demo_login(request):
    """
    Logs in the read-only 'demo' user without password authentication.
    Ensures data is always fresh (current month).
    """
    # Clear messages
    list(messages.get_messages(request))

    try:
        user = User.objects.get(username='demo')
        
        # Check if data is stale (i.e. not from this month)
        # We check the latest expense. If no expenses or old date, refresh.
        last_expense = Expense.objects.filter(user=user).order_by('-date').first()
        is_stale = False
        
        if not last_expense:
            is_stale = True
        else:
            today = date.today()
            if last_expense.date.month != today.month or last_expense.date.year != today.year:
                is_stale = True
        
        if is_stale:
            # Data is old, refresh it
            call_command('setup_demo_user')
            # Refetch the new user object since the old one might have been deleted/recreated
            user = User.objects.get(username='demo')

    except User.DoesNotExist:
        # User doesn't exist, create it
        call_command('setup_demo_user')
        user = User.objects.get(username='demo')

    # Manually set the backend to allow login without authentication
    login(request, user, backend='django.contrib.auth.backends.ModelBackend')
    messages.success(request, "🚀 Welcome to Demo Mode! Feel free to explore the app.")
    return redirect('home')

def demo_signup(request):
    """
    Logs out the demo user and redirects to the signup page.
    """
    logout(request)
    return redirect('signup')

# --------------------
# Mixins
# --------------------

class RecurringTransactionMixin:
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            self.process_recurring_transactions(request.user)
        return super().dispatch(request, *args, **kwargs)

    def process_recurring_transactions(self, user):
        today = date.today()
        recurring_txs = RecurringTransaction.objects.filter(user=user, is_active=True)
        
        for rt in recurring_txs:
            if not rt.last_processed_date:
                current_date = rt.start_date
            else:
                current_date = rt.get_next_date(rt.last_processed_date, rt.frequency)

            while current_date <= today:
                description = f"{rt.description} (Recurring)"
                if rt.transaction_type == 'EXPENSE':
                    Expense.objects.get_or_create(
                        user=user,
                        date=current_date,
                        amount=rt.amount,
                        category=rt.category or 'Uncategorized',
                        defaults={
                            'description': description,
                            'payment_method': rt.payment_method
                        }
                    )
                else:
                    Income.objects.get_or_create(
                        user=user,
                        date=current_date,
                        amount=rt.amount,
                        source=rt.source or 'Other',
                        defaults={'description': description}
                    )
                
                rt.last_processed_date = current_date
                rt.save()
                current_date = rt.get_next_date(current_date, rt.frequency)


# Custom signup view to log user in immediately
class SignUpView(generic.CreateView):
    form_class = CustomSignupForm
    success_url = reverse_lazy('account_login')
    template_name = 'registration/signup.html'

class LandingPageView(TemplateView):
    template_name = 'landing.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return redirect('home')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        plans = SubscriptionPlan.objects.filter(is_active=True)
        context['plans'] = {p.tier: p for p in plans}
        return context

class SettingsHomeView(LoginRequiredMixin, TemplateView):
    template_name = 'expenses/settings_home.html'

@login_required
def home_view(request):
    """
    Dashboard view with filters and multiple charts.
    """
    # Base QuerySet
    expenses = Expense.objects.filter(user=request.user).order_by('-date')
    
    # Filter Logic
    selected_years = request.GET.getlist('year')
    selected_months = request.GET.getlist('month')
    selected_categories = request.GET.getlist('category')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Remove empty strings from lists
    selected_years = [y for y in selected_years if y]
    selected_months = [m for m in selected_months if m]
    selected_categories = [c for c in selected_categories if c]

    # Date Range takes precedence
    if start_date or end_date:
        if start_date:
            expenses = expenses.filter(date__gte=start_date)
        if end_date:
            expenses = expenses.filter(date__lte=end_date)
        
        # Reset lists for UI clarity since we are in range mode
        selected_years = []
        selected_months = []
        
        trend_title = "Expenses Trend (Custom Range)"
    else:
        # Default to current month/year ONLY on initial land (no params)
        if not request.GET and not (selected_years or selected_months):
            selected_years = [str(datetime.now().year)]
            selected_months = [str(datetime.now().month)]
        
        if selected_years:
            expenses = expenses.filter(date__year__in=selected_years)
        if selected_months:
            expenses = expenses.filter(date__month__in=selected_months)
            
        if len(selected_months) == 1 and len(selected_years) == 1:
            trend_title = f"Daily Expenses for {selected_months[0]}/{selected_years[0]}"
        else:
            trend_title = "Monthly Expenses Trend"

    if selected_categories:
        expenses = expenses.filter(category__in=selected_categories)
        
    # Income Logic (Mirroring Expense Filters)
    incomes = Income.objects.filter(user=request.user)
    if start_date or end_date:
        if start_date:
            incomes = incomes.filter(date__gte=start_date)
        if end_date:
            incomes = incomes.filter(date__lte=end_date)
    else:
        if selected_years:
            incomes = incomes.filter(date__year__in=selected_years)
        if selected_months:
            incomes = incomes.filter(date__month__in=selected_months)
    
    total_income = incomes.aggregate(Sum('amount'))['amount__sum'] or 0
    all_dates = Expense.objects.filter(user=request.user).dates('date', 'year', order='DESC')
    years = sorted(list(set([d.year for d in all_dates] + [datetime.now().year])), reverse=True)
    all_categories = Expense.objects.filter(user=request.user).values_list('category', flat=True).distinct().order_by('category')

    # 1. Category Chart Data (Distribution) & Summary Table
    # We need to fetch raw values and merge them in Python to handle whitespace duplicates
    raw_category_data = expenses.values('category').annotate(total=Sum('amount'))
    
    # Process and merge duplicates
    merged_category_map = {}
    for item in raw_category_data:
        # Strip whitespace to normalize
        cat_name = item['category'].strip()
        amount = float(item['total'])
        
        if cat_name in merged_category_map:
            merged_category_map[cat_name] += amount
        else:
            merged_category_map[cat_name] = amount
            
    # Convert back to list of dicts for template/charts, sorted by total
    # This replaces the DB-ordered queryset with a sorted list
    category_data = [
        {'category': cat, 'total': amt} 
        for cat, amt in merged_category_map.items()
    ]
    category_data.sort(key=lambda x: x['total'], reverse=True)

    # Compute limits and usage per category for chart display
    # Compute limits and usage per category for chart display
    category_limits = []
    # Optimization: Pre-fetch all categories for the user to avoid N+1 queries in the loop
    user_categories = {c.name: c for c in Category.objects.filter(user=request.user)}

    for item in category_data:
        cat_name = item['category']
        cat_obj = user_categories.get(cat_name)
        
        limit = float(cat_obj.limit) if (cat_obj and cat_obj.limit) else None
        
        used_percent = round((item['total'] / limit * 100), 1) if limit else None
        category_limits.append({
            'name': cat_name,
            'total': item['total'],
            'limit': limit,
            'used_percent': used_percent,
        })
    
    categories = [item['category'] for item in category_data]
    category_amounts = [item['total'] for item in category_data]
    
    # 2. Time Trend (Stacked) Data
    
    # Determine Labels (X-Axis)
    # Determine Labels (X-Axis)
    if start_date or end_date:
        # For custom range, if range < 60 days, show daily. Else monthly.
        # Simple heuristic: Always show daily for custom range for now, or let logic decide.
        # Let's stick to: if explicit month selected -> daily. If range -> daily (usually granular).
        trend_qs = expenses.annotate(period=TruncDay('date'))
        date_format = '%d %b'
    elif len(selected_months) == 1 and len(selected_years) == 1:
        # Daily view
        trend_qs = expenses.annotate(period=TruncDay('date'))
        date_format = '%d %b'
    else:
        # Monthly view
        trend_qs = expenses.annotate(period=TruncMonth('date'))
        date_format = '%b %Y'

    # Aggregate by Period AND Category for Stacking
    stacked_data = trend_qs.values('period', 'category').annotate(total=Sum('amount')).order_by('period')
    
    # Process into Chart.js Datasets
    # 1. Get unique sorted periods
    periods = sorted(list(set(item['period'] for item in stacked_data)))
    trend_labels = [p.strftime(date_format) for p in periods]
    
    # 2. Build datasets map: { 'CategoryA': [0, 10, 0...], 'CategoryB': ... }
    # Initialize with zeros for all unique NORMALIZED categories found in expenses
    normalized_all_categories = sorted(list(merged_category_map.keys()))
    dataset_map = { cat: [0] * len(periods) for cat in normalized_all_categories }
    
    for item in stacked_data:
        p_idx = periods.index(item['period'])
        # Strip to match our normalized keys
        cat = item['category'].strip()
        if cat in dataset_map:
            dataset_map[cat][p_idx] += float(item['total']) # Add += in case multiple unstripped cats map to same striped cat in same period
            
    # 3. Convert map to list of dataset objects for Chart.js
    trend_datasets = []
    # Define a color palette (Light Blue, Blue Green, Prussian Blue, Honey Yellow, Orange)
    colors = ['#219EBC', '#023047', '#8ECAE6', '#FFB703', '#0575E6']
    
    for i, (cat, data) in enumerate(dataset_map.items()):
        # Only include non-zero datasets
        if sum(data) > 0:
             trend_datasets.append({
                 'label': cat,
                 'data': data,
                 'backgroundColor': colors[i % len(colors)],
                 'borderRadius': 2
             })

    # 3. Top 5 Expenses
    top_expenses_qs = expenses.order_by('-amount')[:5]
    top_labels = [e.description[:20] + '...' if len(e.description) > 20 else e.description for e in top_expenses_qs]
    top_amounts = [float(e.amount) for e in top_expenses_qs]

    # --- NEW: Income vs Expenses Trend Data ---
    # Re-use the truncation logic determined above
    if start_date or end_date or (len(selected_months) == 1 and len(selected_years) == 1):
        trunc_func = TruncDay
    else:
        trunc_func = TruncMonth
        
    inc_trend = incomes.annotate(period=trunc_func('date')).values('period').annotate(total=Sum('amount')).order_by('period')
    exp_trend = expenses.annotate(period=trunc_func('date')).values('period').annotate(total=Sum('amount')).order_by('period')
    
    # Merge periods
    inc_periods = set(i['period'] for i in inc_trend)
    exp_periods = set(e['period'] for e in exp_trend)
    all_periods_sorted = sorted(list(inc_periods.union(exp_periods)))
    
    ie_labels = [p.strftime(date_format) for p in all_periods_sorted]
    ie_income_data = [float(inc_trend.get(period=p)['total']) if inc_trend.filter(period=p).exists() else 0 for p in all_periods_sorted]
    # Optimization: Use dict lookup instead of filter inside loop
    inc_map = {i['period']: float(i['total']) for i in inc_trend}
    exp_map = {e['period']: float(e['total']) for e in exp_trend}
    
    ie_income_data = [inc_map.get(p, 0.0) for p in all_periods_sorted]
    ie_expense_data = [exp_map.get(p, 0.0) for p in all_periods_sorted]
    ie_savings_data = [inc_map.get(p, 0.0) - exp_map.get(p, 0.0) for p in all_periods_sorted]

    # --- NEW: Payment Method Distribution ---
    raw_payment_data = expenses.values('payment_method').annotate(total=Sum('amount')).order_by('payment_method')
    payment_map = {}
    for item in raw_payment_data:
        pm_name = item['payment_method'] or 'Unknown'
        payment_map[pm_name] = float(item['total'])
    
    # Sort by total desc
    sorted_payment_items = sorted(payment_map.items(), key=lambda x: x[1], reverse=True)
    payment_labels = [item[0] for item in sorted_payment_items]
    payment_data = [item[1] for item in sorted_payment_items]


    # 4. Summary Stats
    total_expenses = expenses.aggregate(Sum('amount'))['amount__sum'] or 0
    transaction_count = expenses.count()
    top_category = category_data[0] if category_data else None
    
    savings = total_income - total_expenses

    # --- NEW: Savings Projection (Linear Extrapolation) ---
    current_date = date.today()
    current_year = current_date.year
    current_month = current_date.month 

    # 1. Calculate YTD Savings (Strictly for current year, regardless of filters)
    ytd_income = Income.objects.filter(user=request.user, date__year=current_year, date__month__lte=current_month).aggregate(Sum('amount'))['amount__sum'] or 0
    ytd_expenses = Expense.objects.filter(user=request.user, date__year=current_year, date__month__lte=current_month).aggregate(Sum('amount'))['amount__sum'] or 0
    ytd_savings = ytd_income - ytd_expenses
    
    projected_savings = 0
    
    # Only project if we have data and positive savings
    if ytd_savings > 0:
        # Avoid division by zero if it's January (month 1)
        # Actually, even in Jan, months_passed is 1. So we are good.
        months_passed = current_month
        avg_monthly_savings = ytd_savings / months_passed
        
        months_remaining = 12 - months_passed
        projected_additional = avg_monthly_savings * months_remaining
        
        projected_savings = ytd_savings + projected_additional
    else:
        # If savings are negative or zero, projection is effectively "0" or "current state"
        # We might handle this in template
        projected_savings = 0

    # Calculate MoM Changes ONLY if exactly one year and one month are selected
    prev_month_data = None
    if len(selected_years) == 1 and len(selected_months) == 1:
        try:
            sel_year = int(selected_years[0])
            sel_month = int(selected_months[0])
            
            # Calculate previous month and year
            if sel_month == 1:
                prev_month = 12
                prev_year = sel_year - 1
            else:
                prev_month = sel_month - 1
                prev_year = sel_year

            prev_expenses = Expense.objects.filter(user=request.user, date__year=prev_year, date__month=prev_month).aggregate(Sum('amount'))['amount__sum'] or 0
            prev_income = Income.objects.filter(user=request.user, date__year=prev_year, date__month=prev_month).aggregate(Sum('amount'))['amount__sum'] or 0
            prev_savings = prev_income - prev_expenses

            def calc_pct(current, previous):
                if previous == 0:
                    return None
                return ((current - previous) / previous) * 100

            prev_month_data = {
                'income_pct': calc_pct(total_income, prev_income),
                'expense_pct': calc_pct(total_expenses, prev_expenses),
                'savings_pct': calc_pct(savings, prev_savings),
            }
            # Add absolute values for template display
            for key in list(prev_month_data.keys()):
                val = prev_month_data[key]
                if val is not None:
                    prev_month_data[f'{key}_abs'] = abs(val)
        except (ValueError, IndexError):
            pass

    # Prepare display labels for the template
    display_year = None
    display_month = None
    
    if len(selected_years) == 1:
        display_year = selected_years[0]
        
    if len(selected_months) == 1:
        try:
            m_idx = int(selected_months[0])
            display_month = calendar.month_name[m_idx]
        except (ValueError, IndexError):
            pass

    # NEW: Calculate Previous/Next Month URLs
    prev_month_url = None
    next_month_url = None

    if len(selected_years) == 1 and len(selected_months) == 1:
        try:
            curr_year = int(selected_years[0])
            curr_month = int(selected_months[0])
            
            # Previous Month
            if curr_month == 1:
                pm = 12
                py = curr_year - 1
            else:
                pm = curr_month - 1
                py = curr_year
            
            # Next Month
            if curr_month == 12:
                nm = 1
                ny = curr_year + 1
            else:
                nm = curr_month + 1
                ny = curr_year

            # Construct Query String (Preserve Categories)
            base_qs = []
            for c in selected_categories:
                base_qs.append(f'category={c}')
            
            qs_prev = base_qs + [f'year={py}', f'month={pm}']
            qs_next = base_qs + [f'year={ny}', f'month={nm}']
            
            prev_month_url = f"{reverse('home')}?{'&'.join(qs_prev)}"
            next_month_url = f"{reverse('home')}?{'&'.join(qs_next)}"
            
        except ValueError:
            pass
    
    # --- Emotional Feedback / Insights Logic (Enhanced) ---
    
    insights = []
    
    # helper for streaks
    def get_monthly_savings_status(u, y, m):
        inc = Income.objects.filter(user=u, date__year=y, date__month=m).aggregate(Sum('amount'))['amount__sum'] or 0
        exp = Expense.objects.filter(user=u, date__year=y, date__month=m).aggregate(Sum('amount'))['amount__sum'] or 0
        return inc > exp

    # Construct date params for deep linking
    date_params = ""
    for y in selected_years:
        date_params += f"&year={y}"
    for m in selected_months:
        date_params += f"&month={m}"

    # helper for category links
    def link_cats(cats):
        links_html = format_html_join(
            mark_safe(', '),
            '<a href="{}" class="alert-link text-decoration-underline">{}</a>',
            ((reverse('expense-list') + f"?category={c}{date_params}", c) for c in cats[:2])
        )
        if len(cats) > 2:
            return format_html('{}, etc.', links_html)
        return links_html

    # 0. Anomaly Detection (Spending Spike)
    # Only if viewing current month (or default view)
    is_current_month_view = False
    now = datetime.now()
    if not request.GET or (len(selected_months) == 1 and str(now.month) in selected_months and str(now.year) in selected_years):
         is_current_month_view = True
    
    if is_current_month_view and total_expenses > 0:
        # Calculate last 3 months average
        last_3_months_total = 0
        months_counted = 0
        for i in range(1, 4):
            # Calculate past month/year
            y = now.year
            m = now.month - i
            while m < 1:
                m += 12
                y -= 1
            
            m_total = Expense.objects.filter(user=request.user, date__year=y, date__month=m).aggregate(Sum('amount'))['amount__sum'] or 0
            if m_total > 0:
                last_3_months_total += m_total
                months_counted += 1
        
        if months_counted > 0:
            avg_past_spend = last_3_months_total / months_counted
            
            # Project current month
            days_in_month = calendar.monthrange(now.year, now.month)[1]
            days_passed = now.day
            if days_passed > 0:
                projected_spend = (float(total_expenses) / days_passed) * days_in_month
                avg_past_spend_float = float(avg_past_spend)
                
                if projected_spend > avg_past_spend_float * 1.25 and float(total_expenses) > 1000: # 25% Higher + Min Threshold
                    pct_higher = int(((projected_spend - avg_past_spend_float) / avg_past_spend_float) * 100)
                    insights.append({
                        'type': 'warning',
                        'icon': 'graph-up-arrow',
                        'title': 'Traffic Alert 🚦',
                        'message': f"You're pacing {pct_higher}% higher than usual. Slow down to stay on track!",
                        'allow_share': False
                    })

    # 1. Budget Warnings (High Priority)

    over_budget_cats = [c['name'] for c in category_limits if c['used_percent'] is not None and c['used_percent'] > 100]
    near_budget_cats = [c['name'] for c in category_limits if c['used_percent'] is not None and 90 <= c['used_percent'] <= 100]
    
    # Check savings rate for "Softener" context
    savings_rate = (savings / total_income * 100) if total_income > 0 else 0
    
    if over_budget_cats:
        cats_str = link_cats(over_budget_cats)
        
        if savings_rate >= 20:
            # Contextualized Warning for High Savers
            msg = format_html("Even strong months have leaks. You crossed limits in {} — catching this keeps you on track.", cats_str)
        else:
            # Standard Coaching Warning - "Warning" type (Yellow) instead of Danger (Red) for empathy
            msg = format_html("⚠️ Budget crossed in {} — let’s rebalance to stay safe.", cats_str)

        insights.append({
            'type': 'warning', # Changed from danger
            'icon': 'exclamation-octagon-fill',
            'title': 'Budget Breached',
            'message': msg,
            'allow_share': False
        })
    elif near_budget_cats:
        cats_str = link_cats(near_budget_cats)
        insights.append({
            'type': 'warning',
            'icon': 'exclamation-triangle-fill',
            'title': 'Approaching Limit',
            'message': format_html("Heads up! You're close to overspending on {}.", cats_str),
            'allow_share': False
        })

    # 2. Wins & Cause-Based Praise (Specific & Celebratory)
    if prev_month_data:
        # Calculate Category Savings (Cause of the win)
        # We need prev month category breakdown
        prev_cat_qs = Expense.objects.filter(user=request.user, date__year=prev_year, date__month=prev_month).values('category').annotate(total=Sum('amount'))
        prev_cat_map = {item['category'].strip(): float(item['total']) for item in prev_cat_qs}
        
        savings_contributors = []
        for cat, curr_total in merged_category_map.items():
            prev_total = prev_cat_map.get(cat, 0)
            if prev_total > curr_total:
                diff = prev_total - curr_total
                if diff > 100: # Threshold to mention
                    savings_contributors.append((cat, diff))
        savings_contributors.sort(key=lambda x: x[1], reverse=True)
        top_savers = [c[0] for c in savings_contributors[:2]]
        
        # Savings Win
        if total_income > 0 and savings > 0:
            savings_rate = (savings / total_income) * 100
            if savings_rate >= 20:
                msg_text = f"You've saved {savings_rate:.0f}% of your income this month."
                share_text = f"I saved {savings_rate:.0f}% of my income this month using TrackMyRupee! 🏆"
                
                if top_savers:
                    cats_link = link_cats(top_savers)
                    msg = format_html("{} You spent less on {} — that's where the magic happened.", msg_text, cats_link)
                else:
                    msg = msg_text

                insights.append({
                    'type': 'success',
                    'icon': 'trophy-fill',
                    'title': 'Super Saver Status! 🏆',
                    'message': msg,
                    'allow_share': True,
                    'share_text': share_text
                })
            elif prev_month_data['savings_pct'] and prev_month_data['savings_pct'] > 0:
                 insights.append({
                    'type': 'success',
                    'icon': 'graph-up-arrow',
                    'title': 'Momentum Building 🚀',
                    'message': f"Your savings grew by {prev_month_data['savings_pct_abs']:.0f}% vs last month. You're getting better at this!",
                    'allow_share': True,
                    'share_text': f"My savings grew by {prev_month_data['savings_pct_abs']:.0f}% this month! 🚀 via TrackMyRupee"
                })
        
        # Expense Control Win (if we haven't already praised savings)
        if len(insights) == 0: 
            if prev_month_data['expense_pct'] and prev_month_data['expense_pct'] < -5:
                 msg_text = f"You've cut spending by {prev_month_data['expense_pct_abs']:.0f}%."
                 share_text = f"I cut my spending by {prev_month_data['expense_pct_abs']:.0f}% this month! 👍 via TrackMyRupee"
                 
                 if top_savers:
                     cats_link = link_cats(top_savers)
                     msg = format_html("{} {} saw the biggest drops.", msg_text, cats_link)
                 else:
                     msg = msg_text
                 
                 insights.append({
                    'type': 'success',
                    'icon': 'check-circle-fill',
                    'title': 'You’re in Control 👍',
                    'message': msg,
                    'allow_share': True,
                    'share_text': share_text
                })

    # 3. Streak & Identity (Reassuring / Habit Forming)
    # Only calculate if current status is good
    if savings > 0 and len(selected_years) == 1 and len(selected_months) == 1:
        streak = 1 # Current month counts
        check_to_go = 5 # check max 5 months back
        curr_y_calc, curr_m_calc = int(selected_years[0]), int(selected_months[0])
        
        for i in range(check_to_go):
            # Go back one month
            if curr_m_calc == 1:
                curr_m_calc = 12
                curr_y_calc -= 1
            else:
                curr_m_calc -= 1
            
            if get_monthly_savings_status(request.user, curr_y_calc, curr_m_calc):
                streak += 1
            else:
                break
        
        if streak > 1:
            insights.append({
                'type': 'info', # Use Info for "Identity/Streak"
                'icon': 'fire',
                'title': 'On a Roll!',
                'message': f"🔥 This is your {streak}th month in a row staying under budget.",
                'allow_share': True,
                'share_text': f"🔥 I've stayed under budget for {streak} months in a row! via TrackMyRupee"
            })

    # 4. Fallback
    if not insights and savings > 0:
        insights.append({
            'type': 'info',
            'icon': 'piggy-bank-fill',
            'title': 'In the Green',
            'message': f"You've saved {savings} so far. Keep it up!",
            'allow_share': False
        })
    elif not insights:
        insights.append({
            'type': 'secondary',
            'icon': 'stars',
            'title': 'Fresh Start',
            'message': "Small steps today lead to big results tomorrow. Let's track some expenses!",
            'allow_share': False
        })

    # Limit to top 2 insights to avoid clutter
    insights = insights[:2]

    # Check for onboarding (True if user has NO data at all)
    has_any_data = Expense.objects.filter(user=request.user).exists() or Income.objects.filter(user=request.user).exists()

    context = {
        'is_new_user': not has_any_data,
        'insights': insights[::-1],
        'total_income': total_income,
        'savings': savings,
        'recent_transactions': expenses.order_by('-date')[:5],
        'categories': categories,
        'category_amounts': category_amounts,
        'category_data': category_data, # Passing full queryset for the summary table
        'category_limits': category_limits,
        'trend_labels': trend_labels,
        'trend_datasets': trend_datasets,
        'trend_title': trend_title,
        'top_labels': top_labels,
        'top_amounts': top_amounts,
        # New Context
        'ie_labels': ie_labels,
        'ie_income_data': ie_income_data,
        'ie_expense_data': ie_expense_data,
        'ie_savings_data': ie_savings_data,
        'payment_labels': payment_labels,
        'payment_data': payment_data,
        'years': years,
        'all_categories': all_categories,
        'selected_years': selected_years,
        'selected_months': selected_months,
        'selected_year': display_year,    # NEW: For template display labels
        'selected_month': display_month,  # NEW: For template display labels
        'selected_categories': selected_categories,
        'months_list': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'total_expenses': total_expenses,
        'transaction_count': transaction_count,
        'top_category': top_category,
        'projected_savings': projected_savings, # NEW
        'start_date': start_date,
        'end_date': end_date,
        'prev_month_data': prev_month_data,
        'prev_month_url': prev_month_url,
        'next_month_url': next_month_url,
        'show_tutorial': not request.user.profile.has_seen_tutorial or request.GET.get('tour') == 'true',
        'has_any_budget': any((c.get('limit') or 0) > 0 for c in category_limits),
    }
    return render(request, 'home.html', context)

@login_required
def complete_tutorial(request):
    if request.method == 'POST':
        profile, created = UserProfile.objects.get_or_create(user=request.user)
        profile.has_seen_tutorial = True
        profile.save()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def upload_view(request):
    """
    Upload view with year selection enforcement.
    """
    
    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        selected_year = int(request.POST.get('year'))
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))
                
                if not rows:
                    continue

                # Search for the header row index
                header_row_index = -1
                header_cols = []
                
                for i, row in enumerate(rows[:10]):
                    if not row: continue
                    row_values = [str(val).strip().title() if val is not None else "" for val in row]
                    if 'Date' in row_values and 'Amount' in row_values and 'Description' in row_values:
                        header_row_index = i
                        header_cols = row_values
                        break
                
                if header_row_index == -1:
                    print(f"Skipping sheet {sheet_name}: Could not find header row.")
                    continue

                # Map column indices
                col_map = {col: idx for idx, col in enumerate(header_cols) if col}
                required_columns = ['Date', 'Amount', 'Description', 'Category']
                
                if not all(col in col_map for col in required_columns):
                    print(f"Skipping sheet {sheet_name}: Missing required columns.")
                    continue

                # Process data rows
                for row_data in rows[header_row_index + 1:]:
                    if not any(row_data): continue # Skip empty rows
                    
                    # Parse date
                    date_val = row_data[col_map['Date']]
                    if date_val is None:
                        continue
                        
                    date_obj = None
                    if isinstance(date_val, str):
                        formats = ['%d %b %Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y', '%d %B %Y', '%d %b', '%d-%b', '%d %B']
                        for fmt in formats:
                            try:
                                parsed_date = datetime.strptime(date_val.strip(), fmt).date()
                                date_obj = parsed_date.replace(year=selected_year)
                                break
                            except ValueError:
                                continue
                        if not date_obj:
                            continue
                    elif isinstance(date_val, (datetime, date)):
                        date_obj = date_val if isinstance(date_val, date) else date_val.date()
                        try:
                            date_obj = date_obj.replace(year=selected_year)
                        except ValueError:
                            date_obj = date_obj.replace(day=28, year=selected_year)
                    else:
                        continue # Unsupported date type

                    # Get other fields
                    amount = row_data[col_map['Amount']]
                    description = row_data[col_map['Description']]
                    category = row_data[col_map['Category']] if 'Category' in col_map else None
                    
                    if amount is None or description is None:
                        continue

                    category_obj = None
                    if category:
                        category_name = str(category).strip()
                        if category_name:
                            category_obj, _ = Category.objects.get_or_create(user=request.user, name=category_name)

                    Expense.objects.get_or_create(
                        user=request.user,
                        date=date_obj,
                        amount=float(amount) if not isinstance(amount, float) else amount,
                        description=str(description),
                        category=category_obj.name if category_obj else "Others"
                    )
            return redirect('home')
        except Exception as e:
            print(f"Error processing file: {e}")
            traceback.print_exc()
            pass

    # Context for year dropdown
    current_year = datetime.now().year
    years = range(current_year, current_year - 5, -1)
    
    return render(request, 'upload.html', {'years': years, 'current_year': current_year})

class ExpenseListView(LoginRequiredMixin, RecurringTransactionMixin, ListView):
    model = Expense
    template_name = 'expenses/expense_list.html'
    context_object_name = 'expenses'
    paginate_by = 20

    def get_queryset(self):
        queryset = Expense.objects.filter(user=self.request.user).prefetch_related('tags').order_by('-date')
        
        # Filtering
        selected_years = self.request.GET.getlist('year')
        selected_months = self.request.GET.getlist('month')
        selected_categories = self.request.GET.getlist('category')
        selected_tags = self.request.GET.getlist('tag')
        search_query = self.request.GET.get('search')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        min_amount = self.request.GET.get('min_amount')
        max_amount = self.request.GET.get('max_amount')

        # Remove empty strings from lists
        selected_years = [y for y in selected_years if y]
        selected_months = [m for m in selected_months if m]
        selected_categories = [c for c in selected_categories if c]
        selected_tags = [t for t in selected_tags if t]
        
        # Date Range Logic (Precedence over Year/Month)
        if start_date or end_date:
            if start_date:
                queryset = queryset.filter(date__gte=start_date)
            if end_date:
                queryset = queryset.filter(date__lte=end_date)
        else:
            # Check if any specific filter is active
            has_active_filters = (
                selected_years or 
                selected_months or 
                search_query or
                selected_tags or
                min_amount or
                max_amount
            )
            
            # If no year/month/search filters, default to current month/year
            if not has_active_filters:
                selected_years = [str(datetime.now().year)]
                selected_months = [str(datetime.now().month)]
            
            if selected_years:
                queryset = queryset.filter(date__year__in=selected_years)
            
            if selected_months:
                queryset = queryset.filter(date__month__in=selected_months)

        if selected_categories:
            queryset = queryset.filter(category__in=selected_categories)
        
        # Filter by Tags
        if selected_tags:
            queryset = queryset.filter(tags__id__in=selected_tags).distinct()
        
        # Filter by Amount Range
        if min_amount:
            try:
                queryset = queryset.filter(amount__gte=float(min_amount))
            except ValueError:
                pass
        if max_amount:
            try:
                queryset = queryset.filter(amount__lte=float(max_amount))
            except ValueError:
                pass
        
        # Filter by Payment Method
        payment_method = self.request.GET.get('payment_method')
        if payment_method:
            queryset = queryset.filter(payment_method=payment_method)

        if search_query:
            queryset = queryset.filter(description__icontains=search_query)
            
        # Sorting
        sort_by = self.request.GET.get('sort')
        if sort_by == 'amount_asc':
            queryset = queryset.order_by('amount')
        elif sort_by == 'amount_desc':
            queryset = queryset.order_by('-amount')
        # Default is already '-date' from line 961, so valid fallback.
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate stats for the filtered queryset
        filtered_queryset = self.object_list
        context['filtered_count'] = filtered_queryset.count()
        context['filtered_amount'] = filtered_queryset.aggregate(Sum('amount'))['amount__sum'] or 0

        # Get unique years and categories for validation
        user_expenses = Expense.objects.filter(user=self.request.user)
        years_dates = user_expenses.dates('date', 'year', order='DESC')
        years = sorted(list(set([d.year for d in years_dates] + [datetime.now().year])), reverse=True)
        # Python-side deduplication to handle whitespace variants (e.g. "Goa" vs "Goa ")
        raw_categories = user_expenses.values_list('category', flat=True)
        categories = sorted(list(set([c.strip() for c in raw_categories if c and c.strip()])), key=str.lower)
        
        context['years'] = years
        context['categories'] = categories
        context['months_list'] = [(i, calendar.month_name[i]) for i in range(1, 13)]
        
        # Determine selected year for UI
        # Determine selected year for UI
        year_param = self.request.GET.get('year')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

        context['start_date'] = start_date
        context['end_date'] = end_date
        
        if start_date or end_date:
            context['selected_years'] = []
            context['selected_months'] = []
            context['selected_categories'] = []
        else:
            selected_years = self.request.GET.getlist('year')
            selected_months = self.request.GET.getlist('month')
            selected_categories = self.request.GET.getlist('category')
            search_query = self.request.GET.get('search')
            
            # Remove empty strings
            selected_years = [y for y in selected_years if y]
            selected_months = [m for m in selected_months if m]
            selected_categories = [c for c in selected_categories if c]

            # Check if any specific filter is active
            has_active_filters = (
                selected_years or 
                selected_months or 
                search_query 
                # (ignoring category here as well to match get_queryset)
            )

            # Mirror default logic from get_queryset
            if not has_active_filters:
                selected_years = [str(datetime.now().year)]
                selected_months = [str(datetime.now().month)]
            
            context['selected_years'] = selected_years
            context['selected_months'] = selected_months
            context['selected_categories'] = selected_categories
        
        # Tags for filtering
        context['tags'] = Tag.objects.filter(user=self.request.user).order_by('name')
        context['selected_tags'] = self.request.GET.getlist('tag')
        
        # Amount range filters
        context['min_amount'] = self.request.GET.get('min_amount', '')
        context['max_amount'] = self.request.GET.get('max_amount', '')
        
        # Filter presets
        context['filter_presets'] = FilterPreset.objects.filter(user=self.request.user).order_by('name')
        
        # Payment methods for multi-select
        context['payment_methods'] = Expense.PAYMENT_OPTIONS
        context['selected_payment_methods'] = self.request.GET.getlist('payment_method')
            
        return context

class ExpenseCreateView(LoginRequiredMixin, generic.TemplateView):
    template_name = 'expenses/expense_form.html'

    def get(self, request, *args, **kwargs):
        # We need to wrap the formset to pass 'user' to the form constructor
        ExpenseFormSet = modelformset_factory(Expense, form=ExpenseForm, extra=1, can_delete=True)
        # Pass user to form kwargs using formset_factory's form_kwargs (requires Django 4.0+)
        # For older Django or modelformset, we might need a custom formset or curry the form.
        # Simpler approach: Use a lambda or partial, but modelformset_factory creates a class.
        
        # Actually, best way for modelformset with custom init args is to override BaseFormSet or manually iterate.
        # But simpler hack: Set the widget choices in the view by iterating forms? No, new forms need it.
        
        # Let's use form_kwargs in the formset initialization if supported.
        # Django 1.9+ supports form_kwargs in formset constructor.
        
        initial_data = [{'date': datetime.now().date()} for _ in range(1)]
        formset = ExpenseFormSet(queryset=Expense.objects.none(), initial=initial_data, form_kwargs={'user': request.user})
        next_url = request.GET.get('next', '')
        return render(request, self.template_name, {'formset': formset, 'next_url': next_url})

    def post(self, request, *args, **kwargs):
        ExpenseFormSet = modelformset_factory(Expense, form=ExpenseForm, extra=1, can_delete=True)
        formset = ExpenseFormSet(request.POST, form_kwargs={'user': request.user})
        if formset.is_valid():
            try:
                instances = formset.save(commit=False)
                for instance in instances:
                    instance.user = request.user
                    instance.save()
                
                next_url = request.POST.get('next') or request.GET.get('next')
                if next_url:
                    return redirect(next_url)
                return redirect('expense-list')
            except IntegrityError as e:
                messages.error(request, f"Duplicate record found! You already have this expense recorded for this date.")
                return render(request, self.template_name, {'formset': formset})
        return render(request, self.template_name, {'formset': formset})

class ExpenseUpdateView(LoginRequiredMixin, generic.UpdateView):
    model = Expense
    form_class = ExpenseForm
    template_name = 'expenses/expense_form.html'
    success_url = reverse_lazy('expense-list')

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except IntegrityError:
            messages.error(self.request, "This expense entry already exists.")
            return self.form_invalid(form)

    def get_queryset(self):
        # Ensure user can only edit their own expenses
        return Expense.objects.filter(user=self.request.user)

class ExpenseBulkDeleteView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        expense_ids = request.POST.getlist('expense_ids')
        if not expense_ids:
            messages.error(request, 'No expenses selected for deletion.')
            return redirect('expense-list')
            
        # Filter by IDs and ensuring they belong to the current user for security
        expenses_to_delete = Expense.objects.filter(id__in=expense_ids, user=request.user)
        deleted_count = expenses_to_delete.count()
        
        if deleted_count > 0:
            expenses_to_delete.delete()
            messages.success(request, f'{deleted_count} expenses deleted successfully.')
        else:
            messages.warning(request, 'No valid expenses found to delete.')
            
        return redirect('expense-list')

class ExpenseDeleteView(LoginRequiredMixin, generic.DeleteView):
    model = Expense
    template_name = 'expenses/expense_confirm_delete.html'
    success_url = reverse_lazy('expense-list')

    def get_queryset(self):
        return Expense.objects.filter(user=self.request.user)
    def get_queryset(self):
        return Expense.objects.filter(user=self.request.user)

class CategoryListView(LoginRequiredMixin, generic.ListView):
    model = Category
    template_name = 'expenses/category_list.html'
    context_object_name = 'categories'
    paginate_by = 10

    def get_queryset(self):
        queryset = Category.objects.filter(user=self.request.user).order_by('name')
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(name__icontains=search_query)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        return context

class CategoryCreateView(LoginRequiredMixin, generic.CreateView):
    model = Category
    fields = ['name', 'limit']
    template_name = 'expenses/category_form.html'
    success_url = reverse_lazy('category-list')

    def form_valid(self, form):
        try:
            form.instance.user = self.request.user
            return super().form_valid(form)
        except IntegrityError:
            messages.error(self.request, "This category already exists.")
            return self.form_invalid(form)

class CategoryUpdateView(LoginRequiredMixin, generic.UpdateView):
    model = Category
    fields = ['name', 'limit']
    template_name = 'expenses/category_form.html'
    success_url = reverse_lazy('category-list')

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

    def get_queryset(self):
        return Category.objects.filter(user=self.request.user)
    
    def form_valid(self, form):
        try:
            # Store old name to update related expenses
            old_name = self.get_object().name
            response = super().form_valid(form)
            new_name = self.object.name
            
            if old_name != new_name:
                Expense.objects.filter(user=self.request.user, category=old_name).update(category=new_name)
                
            return response
        except IntegrityError:
            messages.error(self.request, "This category already exists.")
            return self.form_invalid(form)

class CategoryDeleteView(LoginRequiredMixin, generic.DeleteView):
    model = Category
    template_name = 'expenses/category_confirm_delete.html'
    success_url = reverse_lazy('category-list')

    def get_queryset(self):
        return Category.objects.filter(user=self.request.user)

@login_required
def export_expenses(request):
    """
    Export expenses to CSV based on current filters.
    """
    # Check Limits
    if not request.user.profile.is_plus:
        messages.error(request, "Export is available on Plus and Pro plans.")
        return redirect('pricing')

    expenses = Expense.objects.filter(user=request.user).order_by('-date')

    # Filter Logic
    selected_years = request.GET.getlist('year')
    selected_months = request.GET.getlist('month')
    selected_categories = request.GET.getlist('category')
    search_query = request.GET.get('search')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Remove empty strings
    selected_years = [y for y in selected_years if y]
    selected_months = [m for m in selected_months if m]
    selected_categories = [c for c in selected_categories if c]

    # Date Range Logic (Precedence over Year/Month)
    if start_date or end_date:
        if start_date:
            expenses = expenses.filter(date__gte=start_date)
        if end_date:
            expenses = expenses.filter(date__lte=end_date)
    else:
        # Check if any specific filter is active
        has_active_filters = (
            selected_years or 
            selected_months or 
            search_query
        )
        
        # If no year/month/search filters, default to current month/year
        if not has_active_filters:
            selected_years = [str(datetime.now().year)]
            selected_months = [str(datetime.now().month)]
        
        if selected_years:
            expenses = expenses.filter(date__year__in=selected_years)
        
        if selected_months:
            expenses = expenses.filter(date__month__in=selected_months)

    if selected_categories:
        expenses = expenses.filter(category__in=selected_categories)
    if search_query:
        expenses = expenses.filter(description__icontains=search_query)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="expenses.csv"'

    writer = csv.writer(response)
    writer.writerow(['Date', 'Category', 'Description', 'Amount'])

    for expense in expenses:
        writer.writerow([expense.date, expense.category, expense.description, expense.amount])

    return response

# --------------------
# Income Views
# --------------------

class IncomeListView(LoginRequiredMixin, RecurringTransactionMixin, ListView):
    model = Income
    template_name = 'expenses/income_list.html'
    context_object_name = 'incomes'
    paginate_by = 20

    def get_queryset(self):
        queryset = Income.objects.filter(user=self.request.user).order_by('-date')
        
        # Default dates (Current Year)
        today = timezone.localdate()
        default_start = today.replace(month=1, day=1)
        default_end = today.replace(month=12, day=31)

        # Date Filter
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        source = self.request.GET.get('source')

        # Check if we have ANY filter params. If not, apply default dates.
        if not date_from and not date_to and not source:
             self.date_from = default_start.isoformat()
             self.date_to = default_end.isoformat()
             queryset = queryset.filter(date__gte=default_start, date__lte=default_end)
        else:
            # We have some filters (or user explicitly cleared them? - tricky part about "reset")
            # If user wants to "clear" filters, they usually submit empty strings.
            # But the requirement says "default start date...". Usually implies initial load.
            if date_from:
                queryset = queryset.filter(date__gte=date_from)
                self.date_from = date_from
            else:
                self.date_from = ''
            
            if date_to:
                queryset = queryset.filter(date__lte=date_to)
                self.date_to = date_to
            else:
                self.date_to = ''

        # Source Filter
        if source:
            queryset = queryset.filter(source__icontains=source)
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate stats for the filtered queryset
        filtered_queryset = self.object_list
        context['filtered_count'] = filtered_queryset.count()
        context['filtered_amount'] = filtered_queryset.aggregate(Sum('amount'))['amount__sum'] or 0
        
        context['filter_form'] = {
            'date_from': getattr(self, 'date_from', ''),
            'date_to': getattr(self, 'date_to', ''),
            'source': self.request.GET.get('source', ''),
        }
        return context

class IncomeCreateView(LoginRequiredMixin, generic.CreateView):
    model = Income
    form_class = IncomeForm
    template_name = 'expenses/income_form.html'
    success_url = reverse_lazy('income-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        try:
            form.instance.user = self.request.user
            return super().form_valid(form)
        except IntegrityError:
            messages.error(self.request, "This income entry already exists.")
            return self.form_invalid(form)

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context


class IncomeUpdateView(LoginRequiredMixin, generic.UpdateView):
    model = Income
    form_class = IncomeForm
    template_name = 'expenses/income_form.html'
    success_url = reverse_lazy('income-list')

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

    def get_queryset(self):
        return Income.objects.filter(user=self.request.user)

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except IntegrityError:
            messages.error(self.request, "This income entry already exists.")
            return self.form_invalid(form)

class IncomeDeleteView(LoginRequiredMixin, generic.DeleteView):
    model = Income
    template_name = 'expenses/income_confirm_delete.html'
    success_url = reverse_lazy('income-list')

    def get_queryset(self):
        return Income.objects.filter(user=self.request.user)



class CalendarView(LoginRequiredMixin, RecurringTransactionMixin, TemplateView):
    template_name = 'expenses/calendar.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = datetime.now()
        
        # Get year/month from URL or default to current
        year = self.kwargs.get('year', today.year)
        month = self.kwargs.get('month', today.month)
        
        # Validate year/month
        try:
            year = int(year)
            month = int(month)
            if month < 1 or month > 12:
                raise ValueError
        except ValueError:
            year = today.year
            month = today.month

        # Calculate prev/next month for navigation
        if month == 1:
            prev_month_date = date(year - 1, 12, 1)
        else:
            prev_month_date = date(year, month - 1, 1)
            
        if month == 12:
            next_month_date = date(year + 1, 1, 1)
        else:
            next_month_date = date(year, month + 1, 1)

        # Get search query
        search_query = self.request.GET.get('search', '')

        # Base filters
        expense_filters = Q(user=self.request.user, date__year=year, date__month=month)
        income_filters = Q(user=self.request.user, date__year=year, date__month=month)
        
        if search_query:
            # Filter expenses by description or category
            expense_filters &= (Q(description__icontains=search_query) | Q(category__icontains=search_query))
            # Filter income by source or description
            income_filters &= (Q(source__icontains=search_query) | Q(description__icontains=search_query))

        # Get Expense and Income Data for the month
        expenses = Expense.objects.filter(expense_filters).values('date').annotate(total=Sum('amount'))
        
        incomes = Income.objects.filter(income_filters).values('date').annotate(total=Sum('amount'))
        
        # Map data for easy lookup by day
        # Keys are integers (day of month)
        expense_map = {e['date'].day: e['total'] for e in expenses}
        income_map = {i['date'].day: i['total'] for i in incomes}
        
        # Build Calendar Grid
        cal = calendar.Calendar(firstweekday=6) # Start on Sunday
        month_days = cal.monthdayscalendar(year, month)
        
        # Transform into a list of weeks, where each day is an object
        calendar_data = []
        for week in month_days:
            week_data = []
            for day in week:
                if day == 0:
                    week_data.append(None) # Empty slot
                else:
                    week_data.append({
                        'day': day,
                        'income': income_map.get(day, 0),
                        'expense': expense_map.get(day, 0),
                    })
            calendar_data.append(week_data)
        
        
        # Calculate totals for the month to show net savings
        total_monthly_expense = sum(item['total'] for item in expenses) or 0
        total_monthly_income = sum(item['total'] for item in incomes) or 0
        month_net_savings = total_monthly_income - total_monthly_expense

        context['calendar_data'] = calendar_data
        context['current_year'] = year
        context['current_month'] = month
        context['month_name'] = calendar.month_name[month]
        context['month_net_savings'] = month_net_savings
        context['prev_year'] = prev_month_date.year
        context['prev_month'] = prev_month_date.month
        context['next_year'] = next_month_date.year
        context['next_month'] = next_month_date.month
        context['search_query'] = search_query
        
        return context


class BudgetDashboardView(LoginRequiredMixin, RecurringTransactionMixin, TemplateView):
    template_name = 'expenses/budget_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        today = date.today()
        
        month_param = self.request.GET.get('month')
        year_param = self.request.GET.get('year')
        
        month = int(month_param) if month_param else today.month
        year = int(year_param) if year_param else today.year
        
        # Ensure context variables for filters are correct
        context['current_month'] = month
        context['current_year'] = year
        
        categories = Category.objects.filter(user=user)
        budget_data = []
        
        total_budget = 0
        categorized_spent = 0
        
        # Calculate total spending across ALL expenses for the month
        grand_total_spent = Expense.objects.filter(
            user=user,
            date__year=year,
            date__month=month
        ).aggregate(Total=Sum('amount'))['Total'] or 0

        for category in categories:
            spent = Expense.objects.filter(
                user=user,
                category=category.name,
                date__year=year,
                date__month=month
            ).aggregate(Total=Sum('amount'))['Total'] or 0
            
            percentage = (spent / category.limit * 100) if category.limit and category.limit > 0 else 0
            
            budget_data.append({
                'category': category,
                'spent': spent,
                'limit': category.limit,
                'percentage': min(percentage, 100),
                'actual_percentage': percentage,
                'remaining': (category.limit - spent) if category.limit and spent <= category.limit else 0,
                'over_budget': (spent - category.limit) if category.limit and spent > category.limit else 0
            })
            
            if category.limit:
                total_budget += category.limit
            categorized_spent += spent
            
        context.update({
            'budget_data': budget_data,
            'total_budget': total_budget,
            'total_spent': grand_total_spent,
            'total_remaining': (total_budget - grand_total_spent) if total_budget > grand_total_spent else 0,
            'over_budget_amount': (grand_total_spent - total_budget) if grand_total_spent > total_budget else 0,
            'total_percentage': min((grand_total_spent / total_budget * 100), 100) if total_budget else 0,
            'actual_total_percentage': (grand_total_spent / total_budget * 100) if total_budget else 0,
            'month_name': date(year, month, 1).strftime('%B'),
        })

        # MoM Calculation for Budget Dashboard
        if month == 1:
            prev_month = 12
            prev_year = year - 1
        else:
            prev_month = month - 1
            prev_year = year

        prev_spent = Expense.objects.filter(
            user=user,
            date__year=prev_year,
            date__month=prev_month
        ).aggregate(Total=Sum('amount'))['Total'] or 0

        if prev_spent > 0:
            context['spent_mom_pct'] = ((grand_total_spent - prev_spent) / prev_spent) * 100
            context['spent_mom_pct_abs'] = abs(context['spent_mom_pct'])
        else:
            context['spent_mom_pct'] = None
            context['spent_mom_pct_abs'] = None

        context.update({
            'current_month': month,
            'current_year': year,
            'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
            'years': range(today.year - 2, today.year + 2),
        })
        return context

# --------------------
# Recurring Transaction Views
# --------------------

class RecurringTransactionListView(LoginRequiredMixin, ListView):
    model = RecurringTransaction
    template_name = 'expenses/recurring_transaction_list.html'
    context_object_name = 'recurring_transactions'
    filter_expenses_only = True

    def get_queryset(self):
        queryset = RecurringTransaction.objects.filter(user=self.request.user)
        if self.filter_expenses_only:
            queryset = queryset.filter(transaction_type='EXPENSE')
        queryset = queryset.order_by('-created_at')
        
        # Filter by Category
        categories = self.request.GET.getlist('category')
        if categories:
            queryset = queryset.filter(category__in=categories)
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_transactions = self.object_list
        today = date.today()
        
        # Categories for filter
        user_transactions = RecurringTransaction.objects.filter(user=self.request.user)
        categories = user_transactions.values_list('category', flat=True).distinct().order_by('category')
        # Filter out None/Empty if any
        categories = [c for c in categories if c]
        
        context['categories'] = categories
        context['selected_categories'] = self.request.GET.getlist('category')
        
        # Split into Active and Cancelled
        active_subs = [t for t in all_transactions if t.is_active]
        cancelled_subs = [t for t in all_transactions if not t.is_active]
        
        # Calculate Totals (Monthly & Yearly)
        total_monthly = 0
        total_yearly = 0
        
        for sub in active_subs:
            amount = sub.amount
            if sub.frequency == 'DAILY':
                total_monthly += amount * 30
                total_yearly += amount * 365
            elif sub.frequency == 'WEEKLY':
                total_monthly += amount * 4
                total_yearly += amount * 52
            elif sub.frequency == 'MONTHLY':
                total_monthly += amount
                total_yearly += amount * 12
            elif sub.frequency == 'YEARLY':
                total_monthly += amount / 12
                total_yearly += amount

        # Identify "Renewing Soon" (This Month)
        renewing_soon = []
        renewals_count = 0
        
        # Helper to find next date relative to today
        for sub in active_subs:
            # Calculate next occurrence
            next_date = sub.start_date
            
            # For simpler logic, we reset the year/month to current to check basic interval
            # But for accurate "days until", we need better logic:
            
            if sub.frequency == 'DAILY':
                next_date = today + timedelta(days=1)
            elif sub.frequency == 'WEEKLY':
                # Find days ahead
                days_ahead = (sub.start_date.weekday() - today.weekday()) % 7
                if days_ahead == 0 and today > sub.start_date: # if today is the day, but older start
                     days_ahead = 7
                elif days_ahead == 0 and today == sub.start_date: # exact match today
                     days_ahead = 0
                else: 
                     # If start_date was future, we wait. If past, we find next.
                     # Simplified: just next occurrence of that weekday
                     if days_ahead <= 0: days_ahead += 7
                
                # Correction: Standard logic to find next matching weekday
                days_ahead = (sub.start_date.weekday() - today.weekday()) 
                if days_ahead <= 0: # Target day already happened this week or is today
                    days_ahead += 7
                next_date = today + timedelta(days=days_ahead)
                
            elif sub.frequency == 'MONTHLY':
                # Occurs on sub.start_date.day every month
                # If today.day > start_date.day, it's next month.
                # If today.day <= start_date.day, it's this month.
                try:
                    if today.day > sub.start_date.day:
                        # Next month
                        month = today.month + 1
                        year = today.year
                        if month > 12:
                            month = 1
                            year += 1
                        next_date = date(year, month, sub.start_date.day)
                    else:
                        # This month
                        next_date = date(today.year, today.month, sub.start_date.day)
                except ValueError: 
                    # Handle end of month issues (e.g. 31st) - simplified to 1st of next-next month
                    next_date = (today.replace(day=1) + timedelta(days=32)).replace(day=1)

            elif sub.frequency == 'YEARLY':
                try:
                    this_year_date = date(today.year, sub.start_date.month, sub.start_date.day)
                    if today > this_year_date:
                        next_date = date(today.year + 1, sub.start_date.month, sub.start_date.day)
                    else:
                        next_date = this_year_date
                except ValueError:
                    next_date = date(today.year, 2, 28)

            # Annotate object
            sub.annotated_next_date = next_date
            sub.annotated_days_until = (next_date - today).days
            
            # Determine urgency
            is_renewing = False
            if sub.transaction_type == 'EXPENSE':
                if sub.annotated_days_until <= 30: # Show mostly anything coming up soon
                     is_renewing = True
            
            if is_renewing:
                renewing_soon.append(sub)
                renewals_count += 1
            
            # Sort renewing soon by days until
            renewing_soon.sort(key=lambda x: x.annotated_days_until)

        context.update({
            'active_subs': active_subs,
            'cancelled_subs': cancelled_subs,
            'renewing_soon': renewing_soon,
            'renewals_count': renewals_count,
            'total_monthly_cost': total_monthly,
            'total_yearly_cost': total_yearly,
        })
        return context

class RecurringTransactionManageView(RecurringTransactionListView):
    template_name = 'expenses/recurring_transaction_manage.html'
    filter_expenses_only = False

class RecurringTransactionCreateView(LoginRequiredMixin, CreateView):
    model = RecurringTransaction
    form_class = RecurringTransactionForm
    template_name = 'expenses/recurring_transaction_form.html'
    success_url = reverse_lazy('recurring-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        # Check Limits
        current_count = RecurringTransaction.objects.filter(user=self.request.user, is_active=True).count()
        limit = 0 # Free
        if self.request.user.profile.is_plus:
            limit = 3
        if self.request.user.profile.is_pro:
            limit = float('inf')

        if current_count >= limit:
             messages.error(self.request, f"Recurring Transaction limit reached ({limit}). Please upgrade.")
             return redirect('pricing')
             
        form.instance.user = self.request.user
        messages.success(self.request, 'Recurring transaction created successfully.')
        return super().form_valid(form)

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

class RecurringTransactionUpdateView(LoginRequiredMixin, UpdateView):
    model = RecurringTransaction
    form_class = RecurringTransactionForm
    template_name = 'expenses/recurring_transaction_form.html'
    success_url = reverse_lazy('recurring-list')

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

    def get_queryset(self):
        return RecurringTransaction.objects.filter(user=self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        # Check if is_active changed from True to False
        if 'is_active' in form.changed_data and not form.cleaned_data['is_active']:
            # Cancellation detected
            obj = self.get_object() # get current object
            # Calculate yearly saving
            amount = obj.amount
            if obj.frequency == 'DAILY':
                yearly_saving = amount * 365
            elif obj.frequency == 'WEEKLY':
                yearly_saving = amount * 52
            elif obj.frequency == 'MONTHLY':
                yearly_saving = amount * 12
            else: # YEARLY
                yearly_saving = amount
            
            # Assuming currency symbol is available in request or we use generic. 
            # We can use the profile currency if available, or just a generic prompt. 
            # User request used '₹', but code uses {{ currency_symbol }} in template.
            # We'll try to fetch user currency or default.
            currency = '₹'
            if hasattr(self.request.user, 'userprofile'):
                currency = self.request.user.userprofile.currency
                
            messages.success(self.request, f"You just saved {currency}{yearly_saving:,.0f}/year 🎉")
        else:
            messages.success(self.request, 'Recurring transaction updated successfully.')
            
        return super().form_valid(form)

class RecurringTransactionDeleteView(LoginRequiredMixin, DeleteView):
    model = RecurringTransaction
    template_name = 'expenses/recurring_transaction_confirm_delete.html' # Added template_name for consistency
    success_url = reverse_lazy('recurring-list')

    def get_queryset(self):
        return RecurringTransaction.objects.filter(user=self.request.user)

    def form_valid(self, form):
        # Calculate savings
        obj = self.object
        amount = obj.amount
        if obj.frequency == 'DAILY':
            yearly_saving = amount * 365
        elif obj.frequency == 'WEEKLY':
            yearly_saving = amount * 52
        elif obj.frequency == 'MONTHLY':
            yearly_saving = amount * 12
        else: # YEARLY
            yearly_saving = amount
            
        currency = '₹'
        if hasattr(self.request.user, 'userprofile'):
            currency = self.request.user.userprofile.currency
            
        messages.success(self.request, f"You just saved {currency}{yearly_saving:,.0f}/year 🎉")
        return super().form_valid(form)

class AccountDeleteView(LoginRequiredMixin, DeleteView):
    model = User
    success_url = reverse_lazy('landing')
    template_name = 'expenses/account_confirm_delete.html'

    def get_object(self, queryset=None):
        return self.request.user

    def form_valid(self, form):
        user = self.get_object()
        logout(self.request) # Log out before deleting
        user.delete()
        messages.success(self.request, "Your account has been deleted successfully.")
        return redirect(self.success_url)

class CurrencyUpdateView(LoginRequiredMixin, UpdateView):
    model = UserProfile
    fields = ['currency']
    template_name = 'expenses/currency_settings.html'
    success_url = reverse_lazy('currency-settings')

    def get_object(self, queryset=None):
        profile, created = UserProfile.objects.get_or_create(user=self.request.user)
        return profile

    def form_valid(self, form):
        messages.success(self.request, 'Currency preference updated successfully.')
        return super().form_valid(form)

class ProfileUpdateView(LoginRequiredMixin, UpdateView):
    model = User
    form_class = ProfileUpdateForm
    template_name = 'expenses/profile_settings.html'
    success_url = reverse_lazy('profile-settings')

    def get_object(self):
        return self.request.user

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Profile Settings'
        context['is_social_user'] = SocialAccount.objects.filter(user=self.request.user).exists()
        return context

    def form_valid(self, form):
        messages.success(self.request, "Profile updated successfully.")
        return super().form_valid(form)

def demo_login(request):
    """
    Logs in the read-only 'demo' user without password authentication.
    """
    # Clear any existing messages (e.g. from previous logout)
    list(messages.get_messages(request))

    try:
        user = User.objects.get(username='demo')
        # Manually set the backend to allow login without authentication
        login(request, user, backend='django.contrib.auth.backends.ModelBackend')
        messages.success(request, "🚀 Welcome to Demo Mode! Feel free to explore the app.")
        return redirect('home')
    except User.DoesNotExist:
        messages.error(request, "Demo user not setup. Please contact admin.")
        return redirect('account_login')

class PricingView(TemplateView):
    template_name = 'expenses/pricing.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['RAZORPAY_KEY_ID'] = settings.RAZORPAY_KEY_ID
        plans = SubscriptionPlan.objects.filter(is_active=True)
        context['plans'] = {p.tier: p for p in plans}
        return context

def ping(request):
    return HttpResponse("Pong", status=200)

class ContactView(View):
    template_name = 'contact.html'
    
    # Spam protection settings
    RATE_LIMIT_HOURLY = 3
    RATE_LIMIT_DAILY = 10
    MIN_MESSAGE_LENGTH = 10
    
    # Common spam patterns
    SPAM_KEYWORDS = [
        'precio', 'price check', 'buy now', 'click here', 'earn money',
        'viagra', 'casino', 'lottery', 'prize', 'congratulations',
        'limited offer', 'act now', 'online pharmacy', 'weight loss',
        'make money fast', 'work from home', 'investment opportunity',
        'hola, quería saber', 'please kindly', 'dear friend'
    ]
    
    # Disposable email domains
    DISPOSABLE_DOMAINS = [
        'tempmail.com', 'guerrillamail.com', '10minutemail.com',
        'throwaway.email', 'maildrop.cc', 'mailinator.com',
        'trashmail.com', 'yopmail.com', 'getnada.com'
    ]

    def get(self, request):
        form = ContactForm()
        return render(request, self.template_name, {'form': form})
    
    def _get_client_ip(self, request):
        """Get client IP address from request"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip
    
    def _check_rate_limit(self, ip):
        """Check if IP has exceeded rate limits"""
        
        hourly_key = f'contact_hourly_{ip}'
        daily_key = f'contact_daily_{ip}'
        
        hourly_count = cache.get(hourly_key, 0)
        daily_count = cache.get(daily_key, 0)
        
        if hourly_count >= self.RATE_LIMIT_HOURLY:
            return False, "Too many submissions. Please try again in an hour."
        
        if daily_count >= self.RATE_LIMIT_DAILY:
            return False, "Daily submission limit reached. Please try again tomorrow."
        
        # Increment counters
        cache.set(hourly_key, hourly_count + 1, 3600)  # 1 hour
        cache.set(daily_key, daily_count + 1, 86400)   # 24 hours
        
        return True, None
    
    def _is_spam_content(self, text):
        """Check if text contains spam patterns"""
        text_lower = text.lower()
        
        # Check for URLs (most spam contains links)
        if 'http://' in text_lower or 'https://' in text_lower or 'www.' in text_lower:
            return True, "Messages with URLs are not allowed."
        
        # Check for spam keywords
        for keyword in self.SPAM_KEYWORDS:
            if keyword in text_lower:
                return True, "Your message was flagged as potential spam."
        
        # Check for excessive caps (> 50% uppercase)
        if len(text) > 20:
            caps_count = sum(1 for c in text if c.isupper())
            if caps_count / len(text) > 0.5:
                return True, "Please don't use excessive capitalization."
        
        # Check message length
        if len(text.strip()) < self.MIN_MESSAGE_LENGTH:
            return True, "Please provide a more detailed message."
        
        return False, None
    
    def _is_disposable_email(self, email):
        """Check if email is from a disposable domain"""
        domain = email.split('@')[-1].lower()
        return domain in self.DISPOSABLE_DOMAINS

    def post(self, request):
        form = ContactForm(request.POST)
        
        # This handles validations for all fields including reCAPTCHA (if configured)
        if not form.is_valid():
            messages.error(request, "Please correct the errors below.")
            return render(request, self.template_name, {'form': form})

        # Get cleaned data
        data = form.cleaned_data
        name = data.get('name')
        email = data.get('email')
        subject = data.get('subject')
        message = data.get('message')
        honeypot = data.get('website')
        
        # Layer 1: Honeypot check
        if honeypot:
            # Silently reject spam bots - don't reveal honeypot was triggered
            messages.success(request, "Your message has been sent! We'll get back to you shortly.")
            return redirect('contact')
        
        # Layer 2: Rate limiting
        client_ip = self._get_client_ip(request)
        rate_ok, rate_msg = self._check_rate_limit(client_ip)
        if not rate_ok:
            messages.error(request, rate_msg)
            return render(request, self.template_name, {'form': form})
        
        # Layer 3: Content filtering
        is_spam, spam_msg = self._is_spam_content(subject + ' ' + message)
        if is_spam:
            messages.error(request, spam_msg)
            return render(request, self.template_name, {'form': form})
        
        # Layer 4: Email validation
        if self._is_disposable_email(email):
            messages.error(request, "Please use a permanent email address.")
            return render(request, self.template_name, {'form': form})
        
        # Layer 5: reCAPTCHA verification is handled by form.is_valid()
        
        # All checks passed - send email
        full_message = f"""
        New Contact Form Submission:
        
        Name: {name}
        Email: {email}
        Subject: {subject}
        IP: {client_ip}
        
        Message:
        {message}
        """

        try:
            send_mail(
                subject=f"Contact Form: {subject}",
                message=full_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=['track.my.rupee.app@gmail.com'],
                fail_silently=False,
            )
            messages.success(request, "Your message has been sent! We'll get back to you shortly.")
            return redirect('contact')
        except Exception as e:
            # Log error if possible
            messages.error(request, "Something went wrong. Please try again later.")
            return render(request, self.template_name, {'form': form})

@login_required
def predict_category_view(request):
    """
    AJAX view to predict category based on description.
    """
    if request.method == 'GET':
        description = request.GET.get('description', '').strip()
        if not description:
             return JsonResponse({'category': None})
        
        category = predict_category_ai(description, request.user)
        return JsonResponse({'category': category})

# --------------------
# Notification Views
# --------------------

class NotificationListView(LoginRequiredMixin, ListView):
    model = Notification
    template_name = 'expenses/notification_list.html'
    context_object_name = 'notifications'
    paginate_by = 20

    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user).order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['unread_count'] = Notification.objects.filter(user=self.request.user, is_read=False).count()
        return context

@login_required
def mark_notifications_read(request):
    if request.method == 'POST':
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        messages.success(request, "All notifications marked as read.")
        return redirect('notification-list')
    return redirect('notification-list')

@login_required
def mark_single_notification_read(request, pk):
    try:
        notification = Notification.objects.get(pk=pk, user=request.user)
        notification.is_read = True
        notification.save()
        return JsonResponse({'success': True})
    except Notification.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Notification not found'}, status=404)

@csrf_exempt
def trigger_notifications(request):
    """
    HTTP endpoint to trigger notifications via external cron service (e.g. cron-job.org).
    Secured by a secret key in the URL params: ?secret=YOUR_CRON_SECRET
    """
    secret = request.GET.get('secret')
    
    # Check against dedicated CRON_SECRET
    if not secret or secret != settings.CRON_SECRET:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
        
    try:
        call_command('send_notifications')
        return JsonResponse({'success': True, 'message': 'Notifications triggered successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    if request.method == 'POST':
        notification = get_object_or_404(Notification, pk=pk, user=request.user)
        notification.is_read = True
        notification.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)


class AnalyticsView(LoginRequiredMixin, TemplateView):
    template_name = 'expenses/analytics.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        today = timezone.now().date()
        
        # 1. Monthly Trends (Last 12 Months)
        labels = []
        income_data = []
        expense_data = []
        balance_rate_data = []
        
        # Determine the start date: 1st day of the month 11 months ago
        # If today is Jan 2026, 11 months ago is Feb 2025.
        start_date = (today.replace(day=1) - timedelta(days=365)).replace(day=1)
        
        # Fetch data grouped by Month
        monthly_income = Income.objects.filter(
            user=user, date__gte=start_date
        ).annotate(month=TruncMonth('date')).values('month').annotate(total=Sum('amount')).order_by('month')
        
        monthly_expenses = Expense.objects.filter(
            user=user, date__gte=start_date
        ).annotate(month=TruncMonth('date')).values('month').annotate(total=Sum('amount')).order_by('month')
        
        # Merge data into a map {date: {income: 0, expense: 0}}
        data_map = {}
        
        # Initialize map with all 12 months to ensure 0s for missing months
        # Iterate from start_date to today month by month
        curr = start_date
        while curr <= today:
            d = curr.replace(day=1)
            data_map[d] = {'income': 0, 'expense': 0}
            # Move to next month
            # Carefully handle month increment
            next_month = curr.month + 1
            next_year = curr.year
            if next_month > 12:
                next_month = 1
                next_year += 1
            curr = date(next_year, next_month, 1)

        # Fill with DB data
        # Fill with DB data
        for item in monthly_income:
            if item['month']:
                d = item['month']
                if isinstance(d, datetime):
                    d = d.date()
                d = d.replace(day=1)
                if d in data_map:
                    data_map[d]['income'] = float(item['total'])
                
        for item in monthly_expenses:
             if item['month']:
                d = item['month']
                if isinstance(d, datetime):
                    d = d.date()
                d = d.replace(day=1)
                if d in data_map:
                    data_map[d]['expense'] = float(item['total'])
                
        # Sort and prepare lists
        sorted_keys = sorted(data_map.keys())
        # Limit to last 12 months if while loop went over
        sorted_keys = sorted_keys[-12:]
        
        for k in sorted_keys:
            labels.append(k.strftime('%b %Y'))
            inc = data_map[k]['income']
            exp = data_map[k]['expense']
            income_data.append(inc)
            expense_data.append(exp)
            
            # Balance Rate = (Income - Expense) / Income * 100
            if inc > 0:
                rate = ((inc - exp) / inc) * 100
            else:
                rate = 0
            balance_rate_data.append(round(rate, 1))

        context['chart_labels'] = labels
        context['income_data'] = income_data
        context['expense_data'] = expense_data
        context['balance_rate_data'] = balance_rate_data
        
        # 2. Category Breakdown (Current Year)
        current_year = today.year
        category_stats = Expense.objects.filter(
            user=user, date__year=current_year
        ).values('category').annotate(total=Sum('amount')).order_by('-total')
        
        cat_labels = [x['category'] for x in category_stats]
        cat_data = [float(x['total']) for x in category_stats]
        
        context['cat_labels'] = cat_labels
        context['cat_data'] = cat_data
        
        # 3. Key Metrics (YTD)
        # Recalculate based on DB (more accurate than summing chart data if chart is limited)
        # Use date__lte=today to ensure we don't include future recurring entries or future dates
        ytd_income_agg = Income.objects.filter(user=user, date__year=current_year, date__lte=today).aggregate(Sum('amount'))['amount__sum'] or 0
        ytd_expense_agg = Expense.objects.filter(user=user, date__year=current_year, date__lte=today).aggregate(Sum('amount'))['amount__sum'] or 0
        
        context['total_income_ytd'] = ytd_income_agg
        context['total_expense_ytd'] = ytd_expense_agg
        context['total_balance_ytd'] = ytd_income_agg - ytd_expense_agg
        
        if ytd_income_agg > 0:
            context['avg_balance_rate'] = round(((ytd_income_agg - ytd_expense_agg) / ytd_income_agg) * 100, 1)
        else:
            context['avg_balance_rate'] = 0
        
        # 4. Investment/SIP Data for Net Worth
        from .models import SIPInvestment
        active_sips = SIPInvestment.objects.filter(user=user, is_active=True)
        
        total_invested = sum(sip.total_invested for sip in active_sips)
        current_investment_value = sum(sip.current_value for sip in active_sips)
        investment_returns = current_investment_value - total_invested
        
        context['total_invested'] = total_invested
        context['current_investment_value'] = current_investment_value
        context['investment_returns'] = investment_returns
        context['investment_returns_percentage'] = round((investment_returns / total_invested * 100), 2) if total_invested > 0 else 0
        context['active_sips_count'] = active_sips.count()
        
        # Net Worth = Savings (YTD Balance) + Investment Returns
        context['net_worth'] = context['total_balance_ytd'] + investment_returns
            
        return context


# ============================================
# SIP Investment Views
# ============================================

class SIPListView(LoginRequiredMixin, ListView):
    model = SIPInvestment
    template_name = 'expenses/sip_list.html'
    context_object_name = 'sips'
    paginate_by = 20

    def get_queryset(self):
        from .models import SIPInvestment
        return SIPInvestment.objects.filter(user=self.request.user).order_by('-is_active', '-start_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import SIPInvestment
        
        all_sips = SIPInvestment.objects.filter(user=self.request.user)
        active_sips = all_sips.filter(is_active=True)
        paused_sips = all_sips.filter(is_active=False)
        
        context['active_sips'] = active_sips
        context['paused_sips'] = paused_sips
        context['active_count'] = active_sips.count()
        context['paused_count'] = paused_sips.count()
        
        # Calculate totals from actual expenses
        context['total_paid'] = sum(sip.total_paid for sip in all_sips)
        context['installments_count'] = sum(sip.installments_count for sip in all_sips)
        
        # Count due SIPs
        context['due_count'] = sum(1 for sip in active_sips if sip.is_due)
        
        # Monthly SIP commitment
        monthly_commitment = sum(
            sip.amount_per_installment for sip in active_sips 
            if sip.frequency == 'MONTHLY'
        )
        weekly_commitment = sum(
            sip.amount_per_installment * 4 for sip in active_sips 
            if sip.frequency == 'WEEKLY'
        )
        quarterly_commitment = sum(
            sip.amount_per_installment / 3 for sip in active_sips 
            if sip.frequency == 'QUARTERLY'
        )
        context['monthly_commitment'] = monthly_commitment + weekly_commitment + quarterly_commitment
        
        context['can_create'] = True
        
        return context


@login_required
def pay_sip(request, pk):
    """Mark an SIP installment as paid - creates an expense entry"""
    from .models import SIPInvestment, Expense
    from datetime import date
    
    sip = get_object_or_404(SIPInvestment, pk=pk, user=request.user)
    
    if request.method == 'POST':
        # Create expense for this SIP payment
        expense = Expense.objects.create(
            user=request.user,
            date=date.today(),
            amount=sip.amount_per_installment,
            description=f"SIP: {sip.fund_name}",
            category=sip.category.name if sip.category else "Investments",
            payment_method='NetBanking',
            sip=sip
        )
        
        # Update last_paid_date
        sip.last_paid_date = date.today()
        sip.save()
        
        messages.success(request, f'SIP payment of ₹{sip.amount_per_installment} recorded for {sip.fund_name}')
        return redirect('sip-list')
    
    return redirect('sip-list')


class SIPCreateView(LoginRequiredMixin, CreateView):
    model = SIPInvestment
    form_class = SIPForm
    template_name = 'expenses/sip_form.html'
    success_url = reverse_lazy('sip-list')

    def dispatch(self, request, *args, **kwargs):
        # Only Pro users can create SIPs (temporarily disabled for testing)
        # if not request.user.profile.is_pro:
        #     messages.error(request, 'SIP Tracking is a Pro feature. Please upgrade to access it.')
        #     return redirect('pricing')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.user = self.request.user
        messages.success(self.request, f'SIP "{form.instance.fund_name}" added successfully!')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add New SIP'
        context['button_text'] = 'Add SIP'
        return context


class SIPUpdateView(LoginRequiredMixin, UpdateView):
    model = SIPInvestment
    form_class = SIPForm
    template_name = 'expenses/sip_form.html'
    success_url = reverse_lazy('sip-list')

    def get_queryset(self):
        from .models import SIPInvestment
        return SIPInvestment.objects.filter(user=self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, f'SIP "{form.instance.fund_name}" updated successfully!')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Update SIP'
        context['button_text'] = 'Save Changes'
        return context


class SIPDeleteView(LoginRequiredMixin, DeleteView):
    model = SIPInvestment
    template_name = 'expenses/sip_confirm_delete.html'
    success_url = reverse_lazy('sip-list')

    def get_queryset(self):
        from .models import SIPInvestment
        return SIPInvestment.objects.filter(user=self.request.user)

    def form_valid(self, form):
        messages.success(self.request, 'SIP deleted successfully!')
        return super().form_valid(form)


class PortfolioDashboardView(LoginRequiredMixin, TemplateView):
    """Simple SIP summary view - redirects to SIP list"""
    template_name = 'expenses/sip_list.html'

    def get(self, request, *args, **kwargs):
        # Redirect to SIP list since we simplified the feature
        return redirect('sip-list')


# ============================================
# Tag Views
# ============================================

class TagListView(LoginRequiredMixin, ListView):
    model = Tag
    template_name = 'expenses/tag_list.html'
    context_object_name = 'tags'

    def get_queryset(self):
        from .models import Tag
        return Tag.objects.filter(user=self.request.user).order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import Tag
        
        # Add usage count for each tag
        tags_with_count = []
        for tag in context['tags']:
            tags_with_count.append({
                'tag': tag,
                'usage_count': tag.expenses.count()
            })
        context['tags_with_count'] = tags_with_count
        return context


class TagCreateView(LoginRequiredMixin, CreateView):
    model = Tag
    form_class = TagForm
    template_name = 'expenses/tag_form.html'
    success_url = reverse_lazy('tag-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.user = self.request.user
        messages.success(self.request, f'Tag "{form.instance.name}" created successfully!')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Tag'
        context['button_text'] = 'Create'
        return context


class TagUpdateView(LoginRequiredMixin, UpdateView):
    model = Tag
    form_class = TagForm
    template_name = 'expenses/tag_form.html'
    success_url = reverse_lazy('tag-list')

    def get_queryset(self):
        from .models import Tag
        return Tag.objects.filter(user=self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, f'Tag "{form.instance.name}" updated successfully!')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Edit Tag'
        context['button_text'] = 'Save Changes'
        return context


class TagDeleteView(LoginRequiredMixin, DeleteView):
    model = Tag
    template_name = 'expenses/tag_confirm_delete.html'
    success_url = reverse_lazy('tag-list')

    def get_queryset(self):
        from .models import Tag
        return Tag.objects.filter(user=self.request.user)

    def form_valid(self, form):
        messages.success(self.request, 'Tag deleted successfully!')
        return super().form_valid(form)


@login_required
def create_tag_ajax(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            name = data.get('name', '').strip()
            color = data.get('color', 'primary')
            
            if not name:
                return JsonResponse({'success': False, 'error': 'Tag name cannot be empty.'}, status=400)
            
            from .models import Tag
            tag = Tag.objects.create(user=request.user, name=name, color=color)
            return JsonResponse({
                'success': True, 
                'id': tag.id, 
                'name': tag.name,
                'color': tag.color
            })
            
        except IntegrityError:
            return JsonResponse({'success': False, 'error': 'This tag already exists.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
            
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)


# ============================================
# Filter Preset Views
# ============================================

@login_required
def save_filter_preset(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            name = data.get('name', '').strip()
            filter_config = data.get('filter_config', {})
            
            if not name:
                return JsonResponse({'success': False, 'error': 'Preset name cannot be empty.'}, status=400)
            
            from .models import FilterPreset
            
            # Check limits based on tier
            current_count = FilterPreset.objects.filter(user=request.user).count()
            limit = 3  # Free
            if request.user.profile.is_plus:
                limit = 10
            if request.user.profile.is_pro:
                limit = float('inf')
            
            if current_count >= limit:
                return JsonResponse({
                    'success': False, 
                    'error': f'Filter preset limit reached ({int(limit)}). Please upgrade or delete existing presets.'
                }, status=403)
            
            preset, created = FilterPreset.objects.update_or_create(
                user=request.user,
                name=name,
                defaults={'filter_config': filter_config}
            )
            
            return JsonResponse({
                'success': True,
                'id': preset.id,
                'name': preset.name,
                'created': created
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
            
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)


@login_required
def get_filter_presets(request):
    from .models import FilterPreset
    presets = FilterPreset.objects.filter(user=request.user).order_by('name')
    
    data = [{
        'id': p.id,
        'name': p.name,
        'filter_config': p.filter_config
    } for p in presets]
    
    return JsonResponse({'success': True, 'presets': data})


@login_required
def delete_filter_preset(request, pk):
    if request.method == 'POST':
        try:
            from .models import FilterPreset
            preset = get_object_or_404(FilterPreset, pk=pk, user=request.user)
            preset.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)

